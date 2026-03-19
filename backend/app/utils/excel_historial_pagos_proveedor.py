import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _fmt_gs(value):
    return int(value or 0)


def generar_excel_historial_pagos_proveedor(historial, config, fecha_desde=None, fecha_hasta=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Pagos"

    row = 1
    empresa = config.nombre if config and config.nombre else "HESAKA"
    ws.cell(row=row, column=1, value=empresa).font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value="Historial de Pagos a Proveedores").font = Font(bold=True, size=12)
    row += 1

    if fecha_desde or fecha_hasta:
        desde_txt = fecha_desde.strftime("%d/%m/%Y") if fecha_desde else "Inicio"
        hasta_txt = fecha_hasta.strftime("%d/%m/%Y") if fecha_hasta else "Hoy"
        ws.cell(row=row, column=1, value=f"Periodo: {desde_txt} al {hasta_txt}")
        row += 1

    row += 1
    headers = ["Fecha", "Proveedor", "OS", "Factura", "Clientes", "Metodos", "Comprobantes", "Total", "Estado"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
    row += 1

    for item in historial:
        ws.cell(row=row, column=1, value=item.fecha.strftime("%d/%m/%Y") if item.fecha else "-")
        ws.cell(row=row, column=2, value=item.proveedor_nombre or "-")
        ws.cell(row=row, column=3, value=", ".join(item.os_origen or []) or "-")
        ws.cell(row=row, column=4, value=", ".join(item.facturas or []) or "-")
        ws.cell(row=row, column=5, value=", ".join(item.clientes or []) or "-")
        ws.cell(row=row, column=6, value=", ".join(item.metodos or []) or "-")
        ws.cell(row=row, column=7, value=", ".join(item.comprobantes or []) or "-")
        ws.cell(row=row, column=8, value=_fmt_gs(item.total))
        ws.cell(row=row, column=9, value=item.estado or "ACTIVO")
        row += 1

    widths = {
        "A": 14,
        "B": 28,
        "C": 24,
        "D": 24,
        "E": 36,
        "F": 18,
        "G": 24,
        "H": 16,
        "I": 14,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
