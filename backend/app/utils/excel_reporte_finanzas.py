import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _es_movimiento_egreso_para_excel(tipo: str | None) -> bool:
    t = (tipo or "").strip().upper()
    if t in {"INGRESO", "AJUSTE (+)"}:
        return False
    if t in {"EGRESO", "GASTO", "AJUSTE (-)"}:
        return True
    return "EGRESO" in t or "GASTO" in t or "(-)" in (tipo or "")


def generar_excel_reporte_finanzas(
    resumen,
    config,
    fecha_desde=None,
    fecha_hasta=None,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Financiero"

    font_title = Font(name="Arial", size=16, bold=True, color="2C3E50")
    font_subtitle = Font(name="Arial", size=11, italic=True)
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_income = Font(name="Arial", size=10, bold=True, color="15803D")
    font_expense = Font(name="Arial", size=10, bold=True, color="B91C1C")
    font_result_positive = Font(name="Arial", size=10, bold=True, color="1D4ED8")
    font_result_negative = Font(name="Arial", size=10, bold=True, color="B45309")
    fill_header = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    fill_success = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
    fill_income = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_expense = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fill_result_positive = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    fill_result_negative = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    fill_neutral = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_row_income = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    fill_row_expense = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7"),
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    company_name = config.nombre if (config and config.nombre) else "HESAKA"
    ws.merge_cells("A1:H1")
    ws["A1"] = company_name
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:H2")
    ws["A2"] = "REPORTE FINANCIERO"
    ws["A2"].font = Font(name="Arial", size=14, bold=True, color="34495E")
    ws["A2"].alignment = align_center

    if fecha_desde and fecha_hasta:
        period_text = f"Periodo: {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    else:
        period_text = "Periodo: Movimientos financieros"
    ws.merge_cells("A3:H3")
    ws["A3"] = period_text
    ws["A3"].font = font_subtitle
    ws["A3"].alignment = align_center

    ws.merge_cells("A4:H4")
    ws["A4"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A4"].alignment = align_right

    row_idx = 6
    ws.merge_cells(f"A{row_idx}:B{row_idx}")
    ws[f"A{row_idx}"] = "RESUMEN"
    ws[f"A{row_idx}"].font = Font(name="Arial", size=12, bold=True, color="3498DB")
    row_idx += 1

    resumen_data = [
        ("Metrica", "Valor"),
        ("Total Ingresos", resumen.total_ingresos),
        ("Total Egresos", resumen.total_egresos),
        ("Resultado Neto", resumen.resultado_neto),
        ("Margen", f"{resumen.margen:.2f}%"),
        ("Ingresos Caja", resumen.ingresos_caja),
        ("Ingresos Banco", resumen.ingresos_banco),
        ("Egresos Caja", resumen.egresos_caja),
        ("Egresos Banco", resumen.egresos_banco),
        ("Ingreso del Dia", getattr(resumen, "total_cobrado_ventas_con_saldo", 0.0)),
        ("Credito del Dia", getattr(resumen, "cuentas_por_cobrar_dia", 0.0)),
        ("Venta Total del Dia", getattr(resumen, "venta_total_dia", 0.0)),
        ("Cantidad de Ventas del Dia", getattr(resumen, "cantidad_ventas_dia", 0)),
        ("Ventas con Saldo Pendiente", getattr(resumen, "cantidad_ventas_cobrar_dia", 0)),
        ("Saldo Final Total", resumen.saldo_final_total),
        ("Cantidad Movimientos", len(resumen.todos)),
    ]

    for i, (label, value) in enumerate(resumen_data):
        ws[f"A{row_idx}"] = label
        ws[f"B{row_idx}"] = value
        ws[f"A{row_idx}"].border = border
        ws[f"B{row_idx}"].border = border
        if i == 0:
            ws[f"A{row_idx}"].fill = fill_header
            ws[f"B{row_idx}"].fill = fill_header
            ws[f"A{row_idx}"].font = font_header
            ws[f"B{row_idx}"].font = font_header
            ws[f"A{row_idx}"].alignment = align_center
            ws[f"B{row_idx}"].alignment = align_center
        else:
            ws[f"A{row_idx}"].font = font_bold
            ws[f"B{row_idx}"].alignment = align_right
            if i in [1, 2, 3, 5, 6, 7, 8, 9, 10, 11, 14]:
                ws[f"B{row_idx}"].number_format = "#,##0"
            if i == 1:
                ws[f"A{row_idx}"].fill = fill_income
                ws[f"B{row_idx}"].fill = fill_income
                ws[f"A{row_idx}"].font = font_income
                ws[f"B{row_idx}"].font = font_income
            elif i == 2:
                ws[f"A{row_idx}"].fill = fill_expense
                ws[f"B{row_idx}"].fill = fill_expense
                ws[f"A{row_idx}"].font = font_expense
                ws[f"B{row_idx}"].font = font_expense
            elif i == 3:
                resultado_fill = fill_result_positive if resumen.resultado_neto >= 0 else fill_result_negative
                resultado_font = font_result_positive if resumen.resultado_neto >= 0 else font_result_negative
                ws[f"A{row_idx}"].fill = resultado_fill
                ws[f"B{row_idx}"].fill = resultado_fill
                ws[f"A{row_idx}"].font = resultado_font
                ws[f"B{row_idx}"].font = resultado_font
            elif i in [9, 10, 11, 12, 13]:
                ws[f"A{row_idx}"].fill = fill_neutral
                ws[f"B{row_idx}"].fill = fill_neutral
                ws[f"A{row_idx}"].font = font_bold
                ws[f"B{row_idx}"].font = font_bold
        row_idx += 1

    desglose_medios = list(getattr(resumen, "desglose_medios", []) or [])
    if desglose_medios:
        row_idx += 2
        ws.merge_cells(f"A{row_idx}:E{row_idx}")
        ws[f"A{row_idx}"] = "DESGLOSE POR MEDIO"
        ws[f"A{row_idx}"].font = Font(name="Arial", size=12, bold=True, color="0F766E")
        row_idx += 1

        headers_medios = ["Medio", "Ingresos", "Egresos", "Neto", "Movimientos"]
        for idx, header in enumerate(headers_medios, 1):
            cell = ws.cell(row=row_idx, column=idx)
            cell.value = header
            cell.font = font_header
            cell.fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
            cell.border = border
            cell.alignment = align_center
        row_idx += 1

        for item in desglose_medios:
            values = [
                item.get("medio") or "-",
                item.get("ingresos") or 0,
                item.get("egresos") or 0,
                item.get("neto") or 0,
                item.get("cantidad_movimientos") or 0,
            ]
            for idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=idx)
                cell.value = value
                cell.border = border
                cell.fill = fill_neutral
                cell.alignment = align_right if idx in [2, 3, 4] else align_center if idx == 5 else align_left
                if idx in [2, 3, 4]:
                    cell.number_format = "#,##0"
            row_idx += 1

    row_idx += 2
    ws.merge_cells(f"A{row_idx}:H{row_idx}")
    ws[f"A{row_idx}"] = "DETALLE DE MOVIMIENTOS"
    ws[f"A{row_idx}"].font = Font(name="Arial", size=12, bold=True, color="3498DB")
    row_idx += 1

    headers = ["Fecha", "Origen", "Medio", "Banco", "Categoria", "Tipo", "Concepto", "Referencia", "Monto"]
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border
        cell.alignment = align_center
    row_idx += 1

    for mov in resumen.todos:
        es_egreso = _es_movimiento_egreso_para_excel(mov.tipo)
        row_data = [
            mov.fecha.strftime("%d/%m/%Y %H:%M"),
            mov.origen,
            getattr(mov, "medio", "-") or "-",
            mov.banco_nombre or "-",
            mov.categoria or "-",
            mov.tipo,
            mov.concepto or "-",
            mov.referencia or "-",
            mov.monto,
        ]
        for idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=idx)
            cell.value = value
            cell.border = border
            cell.fill = fill_row_expense if es_egreso else fill_row_income
            if idx == 8:
                cell.number_format = "#,##0"
                cell.alignment = align_right
                cell.font = font_expense if es_egreso else font_income
            elif idx in [1, 2, 3, 4, 5, 6]:
                cell.alignment = align_center
                if idx == 6:
                    cell.font = font_expense if es_egreso else font_income
            else:
                cell.alignment = align_left
        row_idx += 1

    widths = {"A": 18, "B": 12, "C": 16, "D": 18, "E": 18, "F": 14, "G": 34, "H": 18, "I": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
