import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _fmt_gs(value):
    return int(value or 0)


def generar_excel_reporte_compras(resumen, compras_data, config, fecha_desde, fecha_hasta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    row = 1
    empresa = config.nombre if config and config.nombre else "HESAKA"
    ws.cell(row=row, column=1, value=empresa).font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value="Reporte de Compras y Proveedores").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value=f"Periodo: {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}")
    row += 2

    headers_resumen = ["Total Comprado", "Total Pagado", "Saldo Pendiente", "Compras Crédito", "Compras Contado", "Compras con OS"]
    values_resumen = [resumen.total_comprado, resumen.total_pagado, resumen.total_pendiente, resumen.total_credito, resumen.total_contado, resumen.total_os]
    for col, header in enumerate(headers_resumen, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        ws.cell(row=row + 1, column=col, value=_fmt_gs(values_resumen[col - 1]))
    row += 4

    ws.cell(row=row, column=1, value="Resumen por proveedor").font = Font(bold=True, size=11)
    row += 1
    provider_headers = ["Proveedor", "Compras", "Total", "Pagado", "Pendiente"]
    for col, header in enumerate(provider_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
    row += 1
    for item in resumen.por_proveedor:
        ws.cell(row=row, column=1, value=item.proveedor_nombre)
        ws.cell(row=row, column=2, value=item.cantidad_compras)
        ws.cell(row=row, column=3, value=_fmt_gs(item.total_comprado))
        ws.cell(row=row, column=4, value=_fmt_gs(item.total_pagado))
        ws.cell(row=row, column=5, value=_fmt_gs(item.saldo_pendiente))
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Detalle de compras").font = Font(bold=True, size=11)
    row += 1
    headers = ["Fecha", "Proveedor", "Ventas", "Clientes", "OS", "Documento", "Condición", "Tipo compra", "Estado", "Entrega", "Total", "Pagado", "Saldo"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
    row += 1

    for compra in compras_data:
        ws.cell(row=row, column=1, value=compra["fecha"].strftime("%d/%m/%Y") if compra.get("fecha") else "-")
        ws.cell(row=row, column=2, value=compra.get("proveedor_nombre") or "-")
        ws.cell(row=row, column=3, value=compra.get("ventas_codigos") or "-")
        ws.cell(row=row, column=4, value=compra.get("clientes") or "-")
        ws.cell(row=row, column=5, value=compra.get("nro_os") or "-")
        ws.cell(row=row, column=6, value=f'{compra.get("tipo_documento") or "-"} {compra.get("nro_factura") or ""}'.strip())
        ws.cell(row=row, column=7, value=compra.get("condicion_pago") or "-")
        ws.cell(row=row, column=8, value=compra.get("tipo_compra") or "-")
        ws.cell(row=row, column=9, value=compra.get("estado") or "-")
        ws.cell(row=row, column=10, value=compra.get("estado_entrega") or "-")
        ws.cell(row=row, column=11, value=_fmt_gs(compra.get("total")))
        ws.cell(row=row, column=12, value=_fmt_gs(compra.get("total_pagado")))
        ws.cell(row=row, column=13, value=_fmt_gs(compra.get("saldo")))
        row += 1

    widths = {
        "A": 14, "B": 28, "C": 18, "D": 28, "E": 18, "F": 22,
        "G": 14, "H": 16, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
