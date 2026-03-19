import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generar_excel_reporte_ventas(ventas_data, config, fecha_desde=None, fecha_hasta=None, 
                                 total_comisiones_referidores=0.0, total_comisiones_bancarias=0.0) -> io.BytesIO:
    """
    Genera un archivo Excel con el reporte de ventas y sus métricas financieras.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    
    # Estilos
    font_title = Font(name='Arial', size=16, bold=True, color="2C3E50")
    font_subtitle = Font(name='Arial', size=12, italic=True)
    font_bold = Font(name='Arial', size=10, bold=True)
    font_normal = Font(name='Arial', size=10)
    
    fill_header = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    fill_header_summary = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    fill_success = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
    
    font_header = Font(name='Arial', size=10, bold=True, color="FFFFFF")
    
    border_thin = Border(
        left=Side(style='thin', color="BDC3C7"),
        right=Side(style='thin', color="BDC3C7"),
        top=Side(style='thin', color="BDC3C7"),
        bottom=Side(style='thin', color="BDC3C7")
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # Encabezado Empresa
    company_name = config.nombre if (config and config.nombre) else "CENTRO OPTICO SANTA FE"
    ws.merge_cells('A1:H1')
    ws['A1'] = company_name
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center
    
    # Titulo de reporte
    ws.merge_cells('A2:H2')
    ws['A2'] = "REPORTE DE VENTAS Y RENTABILIDAD"
    ws['A2'].font = Font(name='Arial', size=14, bold=True, color="34495E")
    ws['A2'].alignment = align_center
    
    # Periodo
    if fecha_desde and fecha_hasta:
        period_text = f"Período: {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    else:
        period_text = "Período: Todas las ventas"
    
    ws.merge_cells('A3:H3')
    ws['A3'] = period_text
    ws['A3'].font = font_subtitle
    ws['A3'].alignment = align_center
    
    gen_date = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells('A4:H4')
    ws['A4'] = gen_date
    ws['A4'].font = font_normal
    ws['A4'].alignment = align_right
    
    row_idx = 6
    
    # Calcular totales
    total_ventas = sum(d['venta'].total for d in ventas_data)
    total_costos = sum(d['costo_total'] for d in ventas_data)
    total_utilidad = sum(d['utilidad_bruta'] for d in ventas_data)
    margen_promedio = (total_utilidad / total_ventas * 100) if total_ventas > 0 else 0
    total_comisiones = total_comisiones_referidores + total_comisiones_bancarias
    utilidad_neta = total_utilidad - total_comisiones
    
    # Resumen Financiero
    ws.merge_cells(f'A{row_idx}:C{row_idx}')
    ws[f'A{row_idx}'] = "RESUMEN FINANCIERO"
    ws[f'A{row_idx}'].font = Font(name='Arial', size=12, bold=True, color="3498DB")
    row_idx += 1
    
    summary_data = [
        ("Métrica", "Valor"),
        ("Total Ventas", total_ventas),
        ("Total Costos", total_costos),
        ("Utilidad Bruta", total_utilidad),
        ("Total Comisiones", total_comisiones),
        ("  - Referidores", total_comisiones_referidores),
        ("  - Bancarias", total_comisiones_bancarias),
        ("Utilidad Neta", utilidad_neta),
        ("Margen Bruto Prom.", f"{margen_promedio:.2f}%"),
        ("Cantidad Ventas", len(ventas_data)),
    ]
    
    for i, (metric, val) in enumerate(summary_data):
        ws[f'A{row_idx}'] = metric
        ws[f'B{row_idx}'] = val
        ws.merge_cells(f'B{row_idx}:C{row_idx}')
        
        # Styles for summary
        if i == 0:
            ws[f'A{row_idx}'].fill = fill_header_summary
            ws[f'A{row_idx}'].font = font_header
            ws[f'B{row_idx}'].fill = fill_header_summary
            ws[f'B{row_idx}'].font = font_header
            ws[f'A{row_idx}'].alignment = align_center
            ws[f'B{row_idx}'].alignment = align_center
        else:
            ws[f'A{row_idx}'].font = font_bold if i in [1, 2, 3, 4, 7] else font_normal
            ws[f'B{row_idx}'].font = font_bold if i in [1, 2, 3, 4, 7] else font_normal
            if i in [1, 2, 3, 4, 5, 6, 7]:  # Currency
                ws[f'B{row_idx}'].number_format = '#,##0'
            ws[f'B{row_idx}'].alignment = align_right
            
            if i == 7: # Utilidad neta highlight
                ws[f'A{row_idx}'].font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
                ws[f'B{row_idx}'].font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
                ws[f'A{row_idx}'].fill = fill_success
                ws[f'B{row_idx}'].fill = fill_success
                
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row_idx}'].border = border_thin
            
        row_idx += 1
        
    row_idx += 2
    
    # Detalle de Ventas
    ws.merge_cells(f'A{row_idx}:H{row_idx}')
    ws[f'A{row_idx}'] = "DETALLE DE VENTAS"
    ws[f'A{row_idx}'].font = Font(name='Arial', size=12, bold=True, color="3498DB")
    row_idx += 1
    
    headers = ['Fecha', 'Factura', 'Cliente', 'Estado', 'Venta', 'Costo', 'Utilid.', 'Margen']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
        
    row_idx += 1
    
    for data in ventas_data:
        venta = data['venta']
        costo_total = data['costo_total']
        utilidad_bruta = data['utilidad_bruta']
        margen_bruto = data['margen_bruto']
        
        cliente_nombre = venta.cliente_rel.nombre if venta.cliente_rel else "-"
        
        row_data = [
            venta.fecha.strftime('%d/%m/%Y'),
            venta.codigo,
            cliente_nombre,
            venta.estado,
            venta.total,
            costo_total,
            utilidad_bruta,
            f"{margen_bruto:.1f}%"
        ]
        
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = val
            cell.font = font_normal
            cell.border = border_thin
            
            if col_num in [1, 2, 4, 8]:
                cell.alignment = align_center
            elif col_num == 3:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                cell.number_format = '#,##0'
                
        row_idx += 1
        
    # Ajustar anchos de columna
    col_widths = {
        'A': 12, 'B': 15, 'C': 35, 'D': 12, 'E': 15, 'F': 15, 'G': 15, 'H': 12
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
