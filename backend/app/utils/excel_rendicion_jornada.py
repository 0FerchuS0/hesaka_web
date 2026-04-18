import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def generar_excel_rendicion_jornada(rendicion, resumen, config):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rendicion"

    font_title = Font(name="Arial", size=16, bold=True, color="1E293B")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_income = Font(name="Arial", size=10, bold=True, color="15803D")
    font_expense = Font(name="Arial", size=10, bold=True, color="B91C1C")
    fill_header = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    fill_soft = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_row_income = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    fill_row_expense = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    company_name = config.nombre if (config and config.nombre) else "HESAKA"
    ws.merge_cells("A1:F1")
    ws["A1"] = company_name
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:F2")
    ws["A2"] = "RENDICION DE JORNADA"
    ws["A2"].font = Font(name="Arial", size=14, bold=True, color="1E293B")
    ws["A2"].alignment = align_center

    ws.merge_cells("A3:F3")
    ws["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].alignment = align_right

    row = 5
    data = [
        ("Campo", "Valor"),
        ("Rendido a", rendicion.rendido_a),
        ("Usuario", rendicion.usuario_nombre or "-"),
        ("Fecha rendicion", rendicion.fecha_hora_rendicion.strftime("%d/%m/%Y %H:%M")),
        ("Monto sugerido", rendicion.monto_sugerido),
        ("Monto rendido", rendicion.monto_rendido),
        ("Diferencia", (rendicion.monto_rendido or 0) - (rendicion.monto_sugerido or 0)),
        ("Observacion", rendicion.observacion or "Sin observacion"),
    ]
    if getattr(rendicion, "fecha_hora_ultima_edicion", None):
        data.extend([
            ("Estado", "EDITADA"),
            ("Fecha original", rendicion.fecha_hora_original.strftime("%d/%m/%Y %H:%M") if rendicion.fecha_hora_original else "-"),
            ("Rendido a original", rendicion.rendido_a_original or "-"),
            ("Monto original", rendicion.monto_rendido_original or 0),
            ("Editada por", rendicion.usuario_ultima_edicion_nombre or "-"),
            ("Ultima edicion", rendicion.fecha_hora_ultima_edicion.strftime("%d/%m/%Y %H:%M")),
            ("Motivo del ajuste", rendicion.motivo_ajuste or "Sin motivo registrado"),
        ])
    for idx, (label, value) in enumerate(data):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border
        if idx == 0:
            ws[f"A{row}"].fill = fill_header
            ws[f"B{row}"].fill = fill_header
            ws[f"A{row}"].font = font_header
            ws[f"B{row}"].font = font_header
            ws[f"A{row}"].alignment = align_center
            ws[f"B{row}"].alignment = align_center
        else:
            ws[f"A{row}"].font = font_bold
            ws[f"B{row}"].alignment = align_left
            ws[f"A{row}"].fill = fill_soft
            ws[f"B{row}"].fill = fill_soft
            if idx in {3, 4, 5}:
                ws[f"B{row}"].number_format = "#,##0"
                ws[f"B{row}"].alignment = align_right
                ws[f"B{row}"].font = font_income if idx in {3, 4} else font_expense
        row += 1

    desglose_medios = list(getattr(resumen, "desglose_medios", []) or [])
    if desglose_medios:
        row += 2
        headers_medios = ["Medio", "Ingresos", "Egresos", "Neto", "Movimientos"]
        for col, header in enumerate(headers_medios, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = font_header
            cell.fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
            cell.border = border
            cell.alignment = align_center
        row += 1

        for item in desglose_medios:
            values = [
                item.get("medio") or "-",
                item.get("ingresos") or 0,
                item.get("egresos") or 0,
                item.get("neto") or 0,
                item.get("cantidad_movimientos") or 0,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                cell.fill = fill_soft
                cell.alignment = align_right if col in [2, 3, 4] else align_center if col == 5 else align_left
                if col in [2, 3, 4]:
                    cell.number_format = "#,##0"
            row += 1

    row += 2
    headers = ["Fecha", "Origen", "Medio", "Categoria", "Concepto", "Referencia", "Monto"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border
        cell.alignment = align_center
    row += 1

    for movimiento in resumen.todos:
        es_ingreso = movimiento.tipo in {"INGRESO", "AJUSTE (+)"}
        values = [
            movimiento.fecha.strftime("%d/%m/%Y %H:%M"),
            movimiento.origen,
            getattr(movimiento, "medio", "-") or "-",
            movimiento.categoria,
            movimiento.concepto or "-",
            movimiento.referencia or "-",
            movimiento.monto,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            cell.fill = fill_row_income if es_ingreso else fill_row_expense
            if col == 7:
                cell.number_format = "#,##0"
                cell.alignment = align_right
                cell.font = font_income if es_ingreso else font_expense
            else:
                cell.alignment = align_left
        row += 1

    widths = {"A": 18, "B": 12, "C": 16, "D": 20, "E": 32, "F": 20, "G": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
