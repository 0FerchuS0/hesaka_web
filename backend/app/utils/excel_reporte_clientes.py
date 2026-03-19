import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def generar_excel_reporte_clientes(clientes, config, buscar=None, referidor_nombre=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    empresa = config.nombre if config and config.nombre else "HESAKA"

    row = 1
    ws.cell(row=row, column=1, value=empresa).font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value="Listado de Clientes").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value=f"Referidor: {referidor_nombre or 'Todos'}")
    row += 1
    ws.cell(row=row, column=1, value=f"Busqueda: {buscar.strip() if buscar and buscar.strip() else 'Todas'}")
    row += 1
    ws.cell(row=row, column=1, value=f"Total: {len(clientes)} cliente(s)")
    row += 2

    headers = ["Nombre", "CI / RUC", "Telefono", "Email", "Direccion", "Referidor", "Registro"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
    row += 1

    for cliente in clientes:
        ws.cell(row=row, column=1, value=cliente.nombre or "-")
        ws.cell(row=row, column=2, value=cliente.ci or "-")
        ws.cell(row=row, column=3, value=cliente.telefono or "-")
        ws.cell(row=row, column=4, value=cliente.email or "-")
        ws.cell(row=row, column=5, value=cliente.direccion or "-")
        ws.cell(row=row, column=6, value=cliente.referidor_rel.nombre if cliente.referidor_rel else "-")
        ws.cell(row=row, column=7, value=cliente.fecha_registro.strftime("%d/%m/%Y") if cliente.fecha_registro else "-")
        row += 1

    widths = {"A": 28, "B": 16, "C": 18, "D": 28, "E": 34, "F": 24, "G": 14}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
