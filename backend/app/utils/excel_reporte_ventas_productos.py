import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _fmt_int(value):
    return int(round(float(value or 0.0)))


def generar_excel_reporte_ventas_productos(resumen, config, fecha_desde, fecha_hasta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas por Productos"

    row = 1
    empresa = config.nombre if config and config.nombre else "HESAKA"
    ws.cell(row=row, column=1, value=empresa).font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value="Reporte de Ventas por Productos").font = Font(bold=True, size=12)
    row += 1

    if fecha_desde and fecha_hasta:
        periodo = f"Periodo: {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    elif fecha_desde:
        periodo = f"Periodo: Desde {fecha_desde.strftime('%d/%m/%Y')}"
    elif fecha_hasta:
        periodo = f"Periodo: Hasta {fecha_hasta.strftime('%d/%m/%Y')}"
    else:
        periodo = "Periodo: Todo el periodo"
    ws.cell(row=row, column=1, value=periodo)
    row += 2

    headers_resumen = [
        "Total Productos",
        "Cantidad Vendida",
        "Total Ingresos",
        "Total Costos",
        "Utilidad Bruta",
        "Margen Bruto Prom (%)",
    ]
    values_resumen = [
        _fmt_int(resumen.total_productos),
        float(resumen.total_cantidad or 0.0),
        _fmt_int(resumen.total_ingresos),
        _fmt_int(resumen.total_costos),
        _fmt_int(resumen.utilidad_bruta_total),
        float(resumen.margen_bruto_promedio or 0.0),
    ]
    for col, header in enumerate(headers_resumen, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        ws.cell(row=row + 1, column=col, value=values_resumen[col - 1])
    row += 4

    headers = [
        "#",
        "Producto",
        "Categoria",
        "Cantidad Vendida",
        "Ingresos Totales",
        "Costos Totales",
        "Utilidad Bruta",
        "Margen Bruto (%)",
        "Precio Promedio",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
    row += 1

    for idx, item in enumerate(resumen.productos or [], start=1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=item.producto_nombre or "-")
        ws.cell(row=row, column=3, value=item.categoria_nombre or "-")
        ws.cell(row=row, column=4, value=float(item.cantidad_vendida or 0.0))
        ws.cell(row=row, column=5, value=_fmt_int(item.ingresos_totales))
        ws.cell(row=row, column=6, value=_fmt_int(item.costos_totales))
        ws.cell(row=row, column=7, value=_fmt_int(item.utilidad_bruta))
        ws.cell(row=row, column=8, value=float(item.margen_bruto or 0.0))
        ws.cell(row=row, column=9, value=_fmt_int(item.precio_promedio))
        row += 1

    widths = {
        "A": 6, "B": 36, "C": 24, "D": 18, "E": 18, "F": 18, "G": 18, "H": 16, "I": 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
